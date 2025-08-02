import openpyxl
import requests
import os
import os.path
from datetime import datetime, timedelta
import calendar
import time
import random
import warnings

def seconds_to_minutes(seconds):
    """Convert seconds to minutes format
    
    Args:
        seconds (float): Time in seconds
        
    Returns:
        float: Time in minutes with 2 decimal precision, or None if input is None
    """
    if seconds is None:
        return None
    
    minutes = seconds / 60
    # Return as a float value with 2 decimal places, not a string
    return round(minutes, 2)

def get_date_range_for_month(month_name, year):
    """
    Get start and end date for a given month name
    
    Args:
        month_name (str): Month name (Jan, Feb, etc.)
        year (int): Year
        
    Returns:
        tuple: (start_date, end_date) in ISO format
    """
    month_map = {
        'Jan': 1, 'Feb': 2, 'Mar': 3, 'Apr': 4, 
        'May': 5, 'Jun': 6, 'Jul': 7, 'Aug': 8,
        'Sep': 9, 'Oct': 10, 'Nov': 11, 'Dec': 12
    }
    
    month_num = month_map.get(month_name, 1)  # Default to January if not found
    
    # Get first day of the month
    first_day = datetime(year, month_num, 1)
    
    # Get last day of the month
    last_day = datetime(year, month_num, calendar.monthrange(year, month_num)[1], 23, 59, 59)
    
    # Format dates in ISO format expected by the API
    start_date = first_day.strftime("%Y-%m-%dT00:00:00.000000Z")
    end_date = last_day.strftime("%Y-%m-%dT00:00:00.000000Z")
    
    return start_date, end_date

def make_api_request(url, headers, payload, mock_api=False, delay=2.5, verify_ssl=True):
    """
    Make API request with rate limiting and retries
    
    Args:
        url (str): API endpoint URL
        headers (dict): Request headers
        payload (dict): Request payload
        mock_api (bool): Whether to use mock data instead of real API call
        delay (float): Delay between requests in seconds
        verify_ssl (bool): Whether to verify SSL certificates
        
    Returns:
        dict: Response data
        
    Raises:
        requests.exceptions.RequestException: If the request fails after retries
    """
    if mock_api:
        # Mock API response for testing
        print("  Using mock API response")
        mean_seconds = random.randint(120, 7200)  # Random value between 2 minutes and 2 hours
        # Create a mock response that matches the real API structure
        return {
            'data': [{
                'mean_seconds_to_first_ack': mean_seconds,
                'total_incident_count': random.randint(5, 20)
            }]
        }
    
    # Real API call with rate limiting and retries
    max_retries = 7  # Increased from 5 to 7
    retry_count = 0
    retry_delay = 3  # Increased initial delay from 2 to 3 seconds
    
    while retry_count < max_retries:
        try:
            # Add a delay between requests to prevent rate limiting
            time.sleep(delay)
            
            print(f"  Making API request to {url}")
            print(f"  Request payload: {payload}")
            
            # Include timeout to prevent hanging requests
            response = requests.post(url, headers=headers, json=payload, timeout=30, verify=verify_ssl)
            
            # Check if we hit rate limiting
            if response.status_code == 429:
                # Get retry-after header or use exponential backoff
                retry_after = int(response.headers.get('Retry-After', retry_delay))
                print(f"  Rate limited (429). Waiting for {retry_after} seconds before retry...")
                time.sleep(retry_after)
                retry_delay = min(retry_delay * 2, 120)  # Exponential backoff, max 120 seconds (increased from 60)
                retry_count += 1
                continue
                
            # Print more info for non-200 responses
            if response.status_code != 200:
                print(f"  Non-200 response: {response.status_code}")
                print(f"  Response headers: {response.headers}")
                try:
                    print(f"  Response content: {response.text[:200]}...")  # Print first 200 chars of response
                except Exception as e:
                    print(f"  Could not print response content: {e}")
            
            response.raise_for_status()  # Raise exception for other 4XX/5XX responses
            return response.json()
            
        except requests.exceptions.Timeout as e:
            print(f"  Request timed out: {e}")
            retry_count += 1
            if retry_count >= max_retries:
                raise requests.exceptions.RequestException(f"Request timed out after {max_retries} retries: {e}")
                
            # Longer delay for timeouts
            wait_time = retry_delay * 2
            print(f"  Retrying in {wait_time:.2f} seconds... (Attempt {retry_count}/{max_retries})")
            time.sleep(wait_time)
            retry_delay = min(retry_delay * 2, 120)
                
        except requests.exceptions.ConnectionError as e:
            print(f"  Connection error: {e}")
            retry_count += 1
            if retry_count >= max_retries:
                raise requests.exceptions.RequestException(f"Connection error after {max_retries} retries: {e}")
            
            # Add jitter to avoid synchronized retries
            jitter = random.uniform(0, 2)
            wait_time = retry_delay + jitter
            print(f"  Connection failed. Retrying in {wait_time:.2f} seconds... (Attempt {retry_count}/{max_retries})")
            time.sleep(wait_time)
            retry_delay = min(retry_delay * 2, 120)  # Increased max backoff
            
        except requests.exceptions.RequestException as e:
            retry_count += 1
            if retry_count >= max_retries:
                print(f"  All retries failed with error: {e}")
                raise  # Re-raise the exception if we've exhausted retries
            
            # Add jitter to avoid synchronized retries
            jitter = random.uniform(0, 2)
            wait_time = retry_delay + jitter
            print(f"  Request failed: {e}. Retrying in {wait_time:.2f} seconds... (Attempt {retry_count}/{max_retries})")
            time.sleep(wait_time)
            retry_delay = min(retry_delay * 2, 120)  # Increased max backoff
    
    raise requests.exceptions.RequestException(f"Failed after {max_retries} retries")


def process_row(row, id_value, name, month_col_idx, url, headers, payload, mock_api=False, delay=2.5, verify_ssl=True):
    """
    Process a single row in the Excel file
    
    Args:
        row (int): Row number
        id_value (str): Escalation policy ID
        name (str): Escalation policy name
        month_col_idx (int): Column index for the month
        url (str): API endpoint URL
        headers (dict): Request headers
        payload (dict): Request payload
        mock_api (bool): Whether to use mock data instead of real API call
        delay (float): Delay between API requests in seconds
        verify_ssl (bool): Whether to verify SSL certificates
        
    Returns:
        float: Minutes value if successful, None otherwise
    """
    try:
        # Make API request
        data = make_api_request(url, headers, payload, mock_api, delay, verify_ssl)
        
        # Extract the mean_seconds_to_first_ack value from the correct location
        mean_seconds = None
        if 'data' in data and len(data['data']) > 0:
            mean_seconds = data['data'][0].get('mean_seconds_to_first_ack')
            print(f"  Found mean_seconds_to_first_ack: {mean_seconds}")
        else:
            print("  No data found in response or empty data array")
        
        # Convert to minutes format and update the sheet
        minutes_value = seconds_to_minutes(mean_seconds)
        return minutes_value
        
    except Exception as e:
        print(f"  Error processing row: {e}")
        import traceback
        traceback.print_exc()  # Print the full stack trace for debugging
        return None


def fetch_and_update_pagerduty_metrics(mock_api=False, month="Jan", force_update=False, 
                             delay=2.5, start_row=None, end_row=None, verify_ssl=True):
    """
    Fetch PagerDuty metrics and update the Excel file
    
    Args:
        mock_api (bool): Whether to use mock data instead of real API calls
        month (str): Month column to update (Jan, Feb, etc.)
        force_update (bool): If True, update even if a value already exists
        delay (float): Delay between API requests in seconds
        start_row (int): Start processing from this row number (1-based, inclusive)
        end_row (int): Stop processing at this row number (1-based, inclusive)
        verify_ssl (bool): Whether to verify SSL certificates
        
    Note:
        Values are stored as actual numbers in Excel with a 2-decimal place format,
        allowing them to be used in calculations and formulas.
    """
    # Load the Excel file - use absolute path or correct relative path
    script_dir = os.path.dirname(os.path.abspath(__file__))
    excel_file = os.path.join(script_dir, '..', 'src', 'MTTA_calc.xlsx')
    
    if not os.path.exists(excel_file):
        print(f"Error: Excel file not found at {excel_file}")
        return
        
    print(f"Loading Excel file from: {excel_file}")
    wb = openpyxl.load_workbook(excel_file)
    sheet = wb.active
    
    # Find column indices
    headers = [sheet.cell(row=1, column=col).value for col in range(1, sheet.max_column + 1)]
    id_col_idx = headers.index('id') + 1  # +1 because openpyxl is 1-indexed
    
    # Find the month column index
    try:
        month_col_idx = headers.index(month) + 1
    except ValueError:
        print(f"Error: Column '{month}' not found in Excel file")
        print(f"Available columns: {headers}")
        return
    
    # Get API token from environment variable (safer than hardcoding)
    api_token = os.environ.get('PAGERDUTY_API_TOKEN')
    if not api_token and not mock_api:
        print("Error: PAGERDUTY_API_TOKEN environment variable not set")
        return
    
    headers = {
        'Authorization': f'Token token={api_token}',
        'Content-Type': 'application/json',
        'Accept': 'application/vnd.pagerduty+json;version=2',
        'User-Agent': 'PagerDuty-MTTA-Analysis/1.0'
    }
    
    # Print API token status (don't print the actual token)
    if api_token:
        print(f"Using PagerDuty API token (first 4 chars: {api_token[:4]}...)")
    else:
        print("No API token found, running in mock mode")
    
    # Determine row range
    first_row = 2  # Skip header row
    last_row = sheet.max_row
    
    if start_row and start_row > first_row:
        first_row = start_row
        
    if end_row and end_row < last_row:
        last_row = end_row
    
    print(f"Processing rows {first_row} to {last_row} (total: {last_row - first_row + 1} rows)")
    
    # Process each row
    for row in range(first_row, last_row + 1):
        id_value = sheet.cell(row=row, column=id_col_idx).value
        if not id_value:
            print(f"Skipping row {row}: No ID value")
            continue
            
        # Check if month column already has a value (unless force_update is True)
        existing_value = sheet.cell(row=row, column=month_col_idx).value
        # Check if the cell has any value (None or empty string means no value)
        if existing_value is not None and existing_value != "" and not force_update:
            print(f"Skipping row {row}: {sheet.cell(row=row, column=1).value} (ID: {id_value}) - {month} value already exists: {existing_value}")
            continue
            
        name = sheet.cell(row=row, column=1).value
        print(f"Processing: {name} (ID: {id_value})")
        
        # Prepare API request
        url = "https://api.pagerduty.com/analytics/metrics/incidents/all"
        
        # Set date range based on month
        start_date, end_date = get_date_range_for_month(month, 2025)
        
        payload = {
            "filters": {
                "escalation_policy_ids": [id_value],
                "created_at_start": start_date,
                "created_at_end": end_date
            }
        }
        
        # Process the row and get the minutes value
        minutes_value = process_row(
            row, id_value, name, month_col_idx, 
            url, headers, payload, mock_api, delay, verify_ssl
        )
        
        if minutes_value is not None:
            # Update the month column with a numeric value
            # When None is passed, openpyxl will leave the cell empty
            cell = sheet.cell(row=row, column=month_col_idx)
            cell.value = minutes_value
            
            # Explicitly set the cell number format to ensure Excel treats it as a number
            # This format represents a number with 2 decimal places
            cell.number_format = '0.00'
            
            print(f"  Updated {month} column: {minutes_value} minutes")
        else:
            print(f"  Failed to update {month} column for {name}")
            # If there's already a value and we failed, keep the old value
            if existing_value:
                print(f"  Keeping existing value: {existing_value}")
            
    # Save the updated Excel file
    wb.save(excel_file)
    print(f"Excel file updated successfully: {excel_file}")


if __name__ == "__main__":
    import sys
    import argparse
    
    # Set up command line arguments
    parser = argparse.ArgumentParser(description="Update PagerDuty metrics in Excel")
    parser.add_argument("--mock", action="store_true", help="Use mock data instead of real API calls")
    parser.add_argument("--month", default="Jan", help="Month column to update (Jan, Feb, etc.)")
    parser.add_argument("--force", action="store_true", help="Force update even if values exist")
    parser.add_argument("--year", type=int, default=2025, help="Year to get data for")
    parser.add_argument("--delay", type=float, default=2.5, help="Delay between API requests in seconds (default: 2.5)")
    parser.add_argument("--start-row", type=int, help="Start processing from this row number")
    parser.add_argument("--end-row", type=int, help="Stop processing at this row number")
    parser.add_argument("--no-verify-ssl", action="store_true", help="Disable SSL certificate verification (use only for testing)")
    
    args = parser.parse_args()
    
    if args.mock:
        print("Running in mock mode - no actual API calls will be made")
    else:
        print(f"Using {args.delay} second delay between API requests")
        
    # If SSL verification is disabled, show a warning
    if args.no_verify_ssl:
        print("\nWARNING: SSL certificate verification is disabled. This is insecure and should only be used for testing.\n")
        import urllib3
        urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)
    
    fetch_and_update_pagerduty_metrics(
        mock_api=args.mock,
        month=args.month,
        force_update=args.force,
        delay=args.delay,
        start_row=args.start_row,
        end_row=args.end_row,
        verify_ssl=not args.no_verify_ssl
    )
